import copy
import os
import re
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def write_excel(excel_out_dir, file_name_company, file_name_channel_set, file_name_time, sheet_list):

    wb = Workbook()

    for sheet_dict in sheet_list:
        # ['sheet_name', 'sheet_title', 'sheet_title_style', 'sheet_data_style', 'sheet_data_list']
        if sheet_dict['sheet_data_list']:
            ws = wb.create_sheet(sheet_dict['sheet_name'])
            ws.append(sheet_dict['sheet_title'])
            for row in sheet_dict['sheet_data_list']:
                ws.append(row)

            # 设置一个字典用于保存列宽数据
            dims = {}
            # 遍历表格数据，获取自适应列宽数据
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                        # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                        # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                        cell_len = len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
            for col, value in dims.items():
                # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
                ws.column_dimensions[get_column_letter(col)].width = value + 2
            # 复制标题样式
            ws_rows = list(ws.iter_rows())
            for index, col in enumerate(ws_rows[0]):
                col._style = copy.copy(sheet_dict['sheet_title_style'][index]['_style'])
                col.font = copy.copy(sheet_dict['sheet_title_style'][index]['font'])
                col.border = copy.copy(sheet_dict['sheet_title_style'][index]['border'])
                col.fill = copy.copy(sheet_dict['sheet_title_style'][index]['fill'])
                col.number_format = copy.copy(sheet_dict['sheet_title_style'][index]['number_format'])
                col.protection = copy.copy(sheet_dict['sheet_title_style'][index]['protection'])
                col.alignment = copy.copy(sheet_dict['sheet_title_style'][index]['alignment'])
            # 复制数据样式
            for ws_row in ws_rows[1:]:
                for index, col in enumerate(ws_row):
                    col._style = copy.copy(sheet_dict['sheet_data_style'][index]['_style'])
                    col.font = copy.copy(sheet_dict['sheet_data_style'][index]['font'])
                    col.border = copy.copy(sheet_dict['sheet_data_style'][index]['border'])
                    col.fill = copy.copy(sheet_dict['sheet_data_style'][index]['fill'])
                    col.number_format = copy.copy(sheet_dict['sheet_data_style'][index]['number_format'])
                    col.protection = copy.copy(sheet_dict['sheet_data_style'][index]['protection'])
                    col.alignment = copy.copy(sheet_dict['sheet_data_style'][index]['alignment'])

    del wb['Sheet']
    # 检查sheet表是否存在某个渠道的数据, 若不存在则文件名中去除此渠道
    final_file_name_channel_set = set()
    for sheet in wb:
        for file_name_channel in file_name_channel_set:
            if sheet.title.upper().__contains__(file_name_channel.upper()):
                final_file_name_channel_set.add(file_name_channel)

    final_file_name_channel_list = list(final_file_name_channel_set)
    final_file_name_channel_list.sort()

    filename = f'{file_name_company}-{"-".join(final_file_name_channel_list)}-{file_name_time}数据.xlsx'
    excel_file = os.path.join(excel_out_dir, filename)
    wb.save(excel_file)

    return True


def load_excel(excel, data_config, company_group_dict):
    # 1. 加载 excel 文件
    work_book = load_workbook(excel)
    sheet = work_book[data_config['sheet']]
    rows = list(sheet.iter_rows(values_only=True))
    # 2 - 表头
    if data_config['special_type'] == '指定列':
        data_title = [col for index, col in enumerate(rows[0]) if data_config['special_col'].__contains__(get_char(index))]
    elif data_config['special_type'] == '排除列':
        data_title = [col for index, col in enumerate(rows[0]) if not data_config['special_col'].__contains__(get_char(index))]
    else:
        data_title = list(rows[0])
    # 3. 表格数据
    data = []
    for row in rows[1:]:
        # 3.1 - 日期
        if data_config['has_time']:
            data_row = [datetime.strptime(row[rows[0].index(data_config['time_col'])], data_config['time_format'])]
        else:
            data_row = [None]
        # 3.2 - 客户名
        group_name = ''
        for col in data_config['customer_col']:
            company = row[rows[0].index(col)].split('-')[0]
            # todo 本公司主体, 排除掉, 本公司给本公司转账不统计
            if company != '欢聚时代文化传媒（北京）有限公司':
                group_name = company
                if company_group_dict.keys().__contains__(company):
                    group_name = company_group_dict[company]

        data_row.append(group_name)

        # 3.3 - 每行数据
        if data_config['special_type'] == '指定列':
            data_row.append([col for index, col in enumerate(row) if data_config['special_col'].__contains__(get_char(index))])
        elif data_config['special_type'] == '排除列':
            data_row.append([col for index, col in enumerate(row) if not data_config['special_col'].__contains__(get_char(index))])
        else:
            data_row.append(list(row))

        # 数据汇总成二维数组
        if group_name != '':
            data.append(data_row)

    # todo 4. 标题样式和数据样式
    style_rows = list(sheet.iter_rows(max_row=2))

    title_style = []
    for style_row in style_rows[0]:
        style_dict = dict()
        style_dict['_style'] = style_row._style
        style_dict['font'] = style_row.font
        style_dict['border'] = style_row.border
        style_dict['fill'] = style_row.fill
        style_dict['number_format'] = style_row.number_format
        style_dict['protection'] = style_row.protection
        style_dict['alignment'] = style_row.alignment
        title_style.append(style_dict)

    data_style = []
    for style_row in style_rows[1]:
        style_dict = dict()
        style_dict['_style'] = style_row._style
        style_dict['font'] = style_row.font
        style_dict['border'] = style_row.border
        style_dict['fill'] = style_row.fill
        style_dict['number_format'] = style_row.number_format
        style_dict['protection'] = style_row.protection
        style_dict['alignment'] = style_row.alignment
        data_style.append(style_dict)

    return data_title, data, title_style, data_style


def get_index(capital):
    """
    大写字母（Excel列头）转索引
    :param capital: 'A' --> 0, 'AA' --> 26
    :return: int
    """
    number = 0
    capital = capital.upper()
    for char in capital:
        number = number * 26 + ord(char) - ord('A') + 1
    return number - 1


def get_char(number):
    """
    索引转大写字母（Excel列头）
    :param number: 0 --> 'A', 26 --> 'AA'
    :return: str
    """
    factor, moder = divmod(number, 26)
    mod_char = chr(moder + 65)
    if factor:
        mod_char = get_char(factor - 1) + mod_char
    return mod_char


def check_file_type(data_dir, file, config):
    # 初步检查模板是否匹配文件, 主要是列字段对应和时间格式
    excel_file = os.path.join(data_dir, file)
    work_book = load_workbook(excel_file)
    # sheet 名称检查
    try:
        sheet = work_book[config['sheet']]
    except Exception:
        return False, f'[{file}] sheet表名配置错误'
    rows = list(sheet.iter_rows(values_only=True, max_row=2))
    # 字段检查
    if config['has_time'] and not rows[0].__contains__(config['time_col']):
        return False, f'[{file}] 不存在时间列/时间列字段名配置错误'
    for customer_col in config['customer_col']:
        if not rows[0].__contains__(customer_col):
            return False, f'[{file}] 客户名配置错误'
    # 时间格式检查
    if config['has_time']:
        try:
            datetime.strptime(rows[1][rows[0].index(config['time_col'])], config['time_format'])
        except Exception:
            return False, f'[{file}] 时间格式配置错误'

    return True, ''
